from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain import Employee, Line, ShiftType, Roster


def export_roster_to_excel(
    roster: Roster,
    employees: List[Employee],
    resolved_line_by_emp: Dict[str, int],
    template_path: Path,
    output_path: Path,
    sheet_name: str = "9 Week Master",
    start_row: int = 11,
    start_col: int = 2,
) -> None:
    """
    Export roster to an Excel workbook matching the operational roster layout.

    Layout assumptions:
    - Columns represent consecutive calendar days
    - Rows are grouped by line
    - Each employee occupies one row under their line
    - Cell values:
        7  -> Day shift
        19 -> Night shift
        '' -> Off
    """

    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    # Group employees by line, preserving input order
    employees_by_line: Dict[int, List[Employee]] = {}
    for e in employees:
        line_id = resolved_line_by_emp.get(e.emp_id)
        if line_id is None:
            continue
        employees_by_line.setdefault(line_id, []).append(e)

    current_row = start_row

    for line_id in sorted(employees_by_line.keys()):
        line_employees = employees_by_line[line_id]

        for idx, emp in enumerate(line_employees):
            row = current_row + (idx * 2)

            # Employee name / ID column
            write_merged_safe(ws, row, 2, emp.display_name())
        current_row += roster.lines[line_id - 1].max_headcount * 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_merged_safe(ws: Worksheet, row: int, col: int, value) -> None:
    """
    Writes value to a cell, resolving merged ranges automatically.
    """
    cell = ws.cell(row=row, column=col)

    # If cell is part of a merged range, write to the top-left cell
    if cell.coordinate in ws.merged_cells:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                tl_cell = ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                )
                tl_cell.value = value
                return
    else:
        cell.value = value
